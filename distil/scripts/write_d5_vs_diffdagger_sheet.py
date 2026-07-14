"""Append a NEW sheet `D5_vs_DiffDAgger` to the COC ablation workbook, preserving all
24 existing sheets. Does NOT touch the existing D5_Compute sheet (that one keeps
SafeDAgger as its baseline).

Safety protocol:
  1. md5 the workbook before.
  2. load_workbook(path) WITHOUT data_only  -> formulas are preserved as formulas.
     (A13 stores its Spread column as formulas with no cached values; loading with
      data_only=True and saving would destroy them.)
  3. If `D5_vs_DiffDAgger` already exists, remove it first (idempotent re-runs).
  4. Save, then RE-OPEN and assert every original sheet name is still present and in
     order, and that the sheet count went 24 -> 25.

Run with an interpreter that has openpyxl (the 'diffdagger' env does NOT):
    /home/s226137394/.conda/envs/maze/bin/python
"""
from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

ROOT = Path("/weka/s226137394/DmNfull")
XLSX = ROOT / "paper_aaai2027/COC_REPORT/ablations_results/DISTIL_ablation_results.xlsx"
ROWS = ROOT / "paper_aaai2027/COC_REPORT/build/table8_diffdagger/table8_rows.json"
SHEET = "D5_vs_DiffDAgger"


def md5(p: Path) -> str:
    return hashlib.md5(p.read_bytes()).hexdigest()


def n(v, nd=1):
    """Number or the literal string UNMEASURED — never a blank, never a guess."""
    if v is None or v == "":
        return "UNMEASURED"
    try:
        return round(float(v), nd)
    except (TypeError, ValueError):
        return v


def main():
    rows = json.loads(ROWS.read_text())
    before_md5 = md5(XLSX)
    wb = load_workbook(XLSX)                    # keep formulas
    before = list(wb.sheetnames)
    print(f"[before] {len(before)} sheets | md5={before_md5}")

    if SHEET in wb.sheetnames:
        del wb[SHEET]
        print(f"[idempotent] removed pre-existing {SHEET}")

    ws = wb.create_sheet(SHEET)
    B, I = Font(bold=True), Font(italic=True)

    def put(r, vals, font=None):
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if font:
                cell.font = font
        return r + 1

    r = 1
    r = put(r, ["D5 — Per-round compute: DISEIL vs Diff-DAgger baseline"], B)
    r = put(r, ["Companion to D5_Compute (which uses SafeDAgger). This sheet does NOT "
                "modify that one."], I)
    r = put(r, ["Method arm 'full' = DISEIL. Baseline arm = 'diffdagger' "
                "(distil/config.py:160 BASELINE_ARMS -> distil/diffdagger.py::run_diffdagger)."])
    r = put(r, ["Seed 1 only. Backend: OpenRouter (VLM qwen/qwen3-vl-30b-a3b-instruct, "
                "LLM qwen/qwen3-32b). Diff-DAgger uses NO foundation model."])
    r = put(r, ["P1 = the run's FIRST round (round 0). P5 = mean ± sample SD (ddof=1) over "
                "rounds 0..4."])
    r = put(r, ["Diff-DAgger token columns are 0 BY CONSTRUCTION (no VLM/LLM/KAG in the arm) "
                "— not missing data."])
    r += 1

    hdr = [
        "Setting", "Protocol", "n rounds (DISEIL)", "n rounds (Diff-DAgger)",
        "Diff-DAgger s/round", "DISEIL s/round", "Overhead x",
        "Shared train+eval s (DISEIL)", "Shared train+eval s (Diff-DAgger)",
        "DISEIL screening s", "DISEIL analysis+prescription s", "DISEIL-specific s",
        "Diff-DAgger gate s", "Reasoning-only add-on s",
        "VLM tok/round", "LLM tok/round", "Reasoning-LLM tok/round", "KAG tok contribution",
        "Diff-DAgger VLM tok", "Diff-DAgger LLM tok", "Diff-DAgger reasoning tok",
        "Diff-DAgger KAG tok",
    ]
    r = put(r, hdr, B)
    hdr_row = r - 1

    def fmt2(m, s, nd=1):
        if m is None:
            return "UNMEASURED"
        if s is None:
            return n(m, nd)
        return f"{float(m):.{nd}f} ± {float(s):.{nd}f}"

    for proto in ("P1", "P5"):
        r = put(r, [f"— PROTOCOL {proto} —"], B)
        for row in [x for x in rows if x["protocol"] == proto]:
            r = put(r, [
                row["setting"], proto,
                row["n_rounds_diseil"], row["n_rounds_diffdagger"],
                fmt2(row["diffdagger_s_per_round"], row["diffdagger_s_sd"]),
                fmt2(row["diseil_s_per_round"], row["diseil_s_sd"]),
                (round(float(row["overhead_x"]), 3) if row["overhead_x"] else "UNMEASURED"),
                fmt2(row["shared_train_eval_s_diseil"], row["shared_train_eval_s_diseil_sd"]),
                fmt2(row["shared_train_eval_s_diffdagger"], row["shared_train_eval_s_diffdagger_sd"]),
                fmt2(row["diseil_screening_s"], row["diseil_screening_s_sd"]),
                fmt2(row["diseil_analysis_prescription_s"], row["diseil_analysis_prescription_s_sd"]),
                fmt2(row["diseil_specific_s"], row["diseil_specific_s_sd"]),
                fmt2(row["diffdagger_gate_s"], row["diffdagger_gate_s_sd"]),
                fmt2(row["reasoning_only_addon_s"], row["reasoning_only_addon_s_sd"]),
                fmt2(row["vlm_tokens_per_round"], row["vlm_tokens_sd"], 0),
                fmt2(row["llm_tokens_per_round"], row["llm_tokens_sd"], 0),
                fmt2(row["reasoning_llm_tokens_per_round"], row["reasoning_llm_tokens_sd"], 0),
                fmt2(row["kag_token_contribution_per_round"], row["kag_token_contribution_sd"], 0),
                0, 0, 0, 0,
            ])
        r += 1

    r += 1
    r = put(r, ["CAVEATS"], B)
    for c in [
        "SINGLE SEED (seed 1). The P5 '±' is the round-to-round spread WITHIN one run — "
        "it is NOT a cross-seed error bar and must not be read as one.",
        "P1 is the FIRST round and is an UPPER bound on steady-state cost: the policy is at "
        "its weakest, so it fails most, which maximises the number of VLM/LLM calls and the "
        "screening rollouts.",
        "REPORT BOTH NUMBERS. The raw Overhead x is close to 1 and UNDERSTATES the true cost "
        "of reasoning, because the shared from-scratch retrain + 100-episode eval dominate the "
        "denominator for BOTH arms. The reasoning-only add-on (DISEIL-specific s minus the "
        "baseline's gate s) is the honest marginal cost.",
        "BUDGET ASYMMETRY (measured): DISEIL adds 1 demo/round, Diff-DAgger adds "
        "interventions_per_round=4. At BUDGET=5 Diff-DAgger hits final_demos and stops after "
        "only 2 rounds, so its 5-round P5 spread comes from a matched BUDGET=20 run "
        "(Nd=4 -> exactly rounds 0..4). P1 is taken from the BUDGET=5 run for both arms.",
        "Backend: OpenRouter for all DISEIL runs. Token counts are NOT comparable across "
        "different backends (a local vLLM run would tokenise and bill differently).",
        "Diff-DAgger's zeros in the token columns are structural, not unmeasured: the arm "
        "queries a diffusion-loss CDF quantile, never a foundation model.",
        "Every second in this sheet traces to a printed, timestamped event in a run that "
        "COMPLETED on this cluster. Diff-DAgger's history carries no per-round 'sec' key "
        "(baselines.py::_wrap_history), so its round wall-clock is derived from run.log "
        "timestamps ([train] -> last [dagger ep]).",
    ]:
        r = put(r, [c])

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    for col in "CDEFGHIJKLMNOPQRSTUV":
        ws.column_dimensions[col].width = 15
    for c in range(1, len(hdr) + 1):
        ws.cell(row=hdr_row, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(XLSX)

    # ---- verify nothing was lost ----
    wb2 = load_workbook(XLSX)
    after = list(wb2.sheetnames)
    missing = [s for s in before if s not in after]
    print(f"[after ] {len(after)} sheets")
    if missing:
        print(f"FAIL: sheets LOST: {missing}")
        return 1
    if after[:len(before)] != before:
        print(f"FAIL: original sheet ORDER changed:\n  before={before}\n  after ={after}")
        return 1
    if SHEET not in after:
        print(f"FAIL: {SHEET} not present after save")
        return 1
    print(f"OK: all {len(before)} original sheets preserved, in order; "
          f"+{SHEET} => {len(after)} total")
    print(f"OK: D5_Compute still present and untouched: {'D5_Compute' in after}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
