"""Emit the final D5_Compute deliverables from distil/results/_compute/d5_merged.json.

  paper_aaai2027/COC_REPORT/build/d5_compute.md
  paper_aaai2027/COC_REPORT/build/d5_compute.csv
  paper_aaai2027/COC_REPORT/ablations_results/DISTIL_ablation_results.xlsx  [D5_Compute only]

The workbook is opened with openpyxl and ONLY the D5_Compute sheet's cells are written;
every other sheet is left untouched and the sheet list is re-verified after saving.
"""
from __future__ import annotations

import csv
import json
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path("/weka/s226137394/DmNfull")
MERGED = ROOT / "distil/results/_compute/d5_merged.json"
BUILD = ROOT / "paper_aaai2027/COC_REPORT/build"
XLSX = ROOT / "paper_aaai2027/COC_REPORT/ablations_results/DISTIL_ablation_results.xlsx"

UNM = "UNMEASURED"


def f(x, n=0, sd=None):
    if x is None:
        return UNM
    s = f"{x:,.{n}f}"
    if sd is not None:
        s += f" ± {sd:,.{n}f}"
    return s


def main():
    M = json.loads(MERGED.read_text())
    rows, jobs = M["rows"], M["jobs"]
    R = {r["label"]: r for r in rows}
    order = ["Door / state", "Push-T / image", "Wipe / image", "Door / image",
             "GridWorld 5x5 / image"]

    # ---------------------------------------------------------------- CSV
    cols = ["setting", "protocol", "rounds_used", "baseline_s_per_round",
            "diseil_s_per_round", "overhead_x", "vlm_tokens_per_round",
            "llm_tokens_per_round", "kag_tokens_per_round",
            "reasoning_llm_tokens_per_round", "diseil_reasoning_only_s_narrow",
            "diseil_reasoning_only_s_wide", "baseline_gate_s_per_round",
            "reasoning_only_addon_s", "shared_train_eval_s", "screen_s", "llm_s",
            "prescribe_s", "total_tokens_per_round", "llm_calls_per_round",
            "kag_tokens_per_call", "kag_pct_of_prompt_tokens", "status",
            "job_diseil", "job_baseline"]
    out = []
    for lab in order:
        r = R[lab]
        jd, jb = jobs[lab]
        for tag in ("first", "mean"):
            b = r[tag]
            # the P5 label states the ACTUAL number of rounds averaged, so it can never
            # over-claim if a run is still in flight
            pname = ("P1: first round" if tag == "first" else
                     f"P5: mean of {b['n_rounds']} LLM-active rounds" if b else
                     "P5: mean over LLM-active rounds")
            if b is None:
                out.append({
                    "setting": lab, "protocol": pname, "rounds_used": UNM,
                    **{c: UNM for c in cols if c not in
                       ("setting", "protocol", "rounds_used", "status",
                        "job_diseil", "job_baseline")},
                    "status": ("UNMEASURED — run at BUDGET=1 by instruction; this setting "
                               "has exactly one round, so no multi-round mean and no "
                               "spread exist."),
                    "job_diseil": jd, "job_baseline": jb})
                continue
            n = b["n_rounds"]
            out.append({
                "setting": lab, "protocol": pname,
                "rounds_used": " ".join(str(x) for x in b["rounds_used"]),
                "baseline_s_per_round": b["baseline_s"],
                "diseil_s_per_round": b["diseil_s"],
                "overhead_x": b["overhead_x"],
                "vlm_tokens_per_round": b["vlm_tokens"],
                "llm_tokens_per_round": b["llm_tokens"],
                "kag_tokens_per_round": b["kag_tokens"],
                "reasoning_llm_tokens_per_round": b["reasoning_tokens"],
                "diseil_reasoning_only_s_narrow": b["reasoning_narrow_s"],
                "diseil_reasoning_only_s_wide": b["reasoning_wide_s"],
                "baseline_gate_s_per_round": b["baseline_gate_s"],
                "reasoning_only_addon_s": b["reasoning_wide_addon_s"],
                "shared_train_eval_s": b["train_eval_shared_s"],
                "screen_s": b["screen_s"], "llm_s": b["llm_s"],
                "prescribe_s": b["prescribe_s"],
                "total_tokens_per_round": b["total_tokens"],
                "llm_calls_per_round": b["calls"],
                "kag_tokens_per_call": b["kag_per_call"],
                "kag_pct_of_prompt_tokens": b["kag_pct_of_prompt"],
                "status": f"measured ({n} round{'s' if n > 1 else ''})",
                "job_diseil": jd, "job_baseline": jb})
    with open(BUILD / "d5_compute.csv", "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r_ in out:
            w.writerow({c: r_.get(c, "") for c in cols})

    # ---------------------------------------------------------------- XLSX
    shutil.copy2(XLSX, XLSX.with_suffix(".xlsx.bak"))
    wb = openpyxl.load_workbook(XLSX)
    before = list(wb.sheetnames)
    ws = wb["D5_Compute"]
    # rows 9..13 hold the five settings, in this order, cols A(Task) B(Obs) C..I
    sheet_order = ["Door / state", "Push-T / image", "Wipe / image", "Door / image",
                   "GridWorld 5x5 / image"]
    for i, lab in enumerate(sheet_order):
        b = R[lab]["first"]
        row = 9 + i
        ws.cell(row, 3).value = b["baseline_s"]
        ws.cell(row, 4).value = b["diseil_s"]
        ws.cell(row, 5).value = b["vlm_tokens"]
        ws.cell(row, 6).value = b["llm_tokens"]
        ws.cell(row, 7).value = b["overhead_x"]
        ws.cell(row, 8).value = b["kag_tokens"]
        ws.cell(row, 9).value = b["reasoning_tokens"]
    ws.cell(14, 1).value = (
        "FILLED. Protocol P1 (first round of the run) — the only protocol available for "
        "all five settings, because Push-T and GridWorld were run at BUDGET=1. Baseline "
        "arm = SafeDAgger. See the two blocks below, and "
        "paper_aaai2027/COC_REPORT/build/d5_compute.md for the full caveats.")

    r = 16
    def put(row, vals, bold=False):
        for j, v in enumerate(vals):
            c = ws.cell(row, 1 + j)
            c.value = v
            if bold:
                c.font = openpyxl.styles.Font(bold=True)

    put(r, ["BLOCK 1 — Protocol P1: FIRST ROUND (apples-to-apples, all five settings)"], True)
    put(r + 1, ["Task", "Obs", "Baseline s/round", "DISTIL s/round", "VLM tokens/round",
                "LLM tokens/round", "Overhead ×", "KAG token contribution",
                "Reasoning LLM tokens/round", "DISTIL reasoning-only s/round (narrow)",
                "DISTIL reasoning-only s/round (wide)", "Baseline gate s/round",
                "Reasoning-only add-on s", "Shared train+eval s (both arms pay)",
                "Round measured"], True)
    for i, lab in enumerate(sheet_order):
        b = R[lab]["first"]
        t, o = lab.split(" / ")
        put(r + 2 + i, [t, o, b["baseline_s"], b["diseil_s"], b["vlm_tokens"],
                        b["llm_tokens"], b["overhead_x"], b["kag_tokens"],
                        b["reasoning_tokens"], b["reasoning_narrow_s"],
                        b["reasoning_wide_s"], b["baseline_gate_s"],
                        b["reasoning_wide_addon_s"], b["train_eval_shared_s"],
                        b["rounds_used"][0]])

    r2 = r + 8
    put(r2, ["BLOCK 2 — Protocol P5: MEAN ± sample SD over the LLM-active rounds "
             "(BUDGET=5 RoboSuite runs only)"], True)
    put(r2 + 1, ["Task", "Obs", "Baseline s/round", "DISTIL s/round", "VLM tokens/round",
                 "LLM tokens/round", "Overhead ×", "KAG token contribution",
                 "Reasoning LLM tokens/round", "DISTIL reasoning-only s/round (narrow)",
                 "DISTIL reasoning-only s/round (wide)", "Baseline gate s/round",
                 "Reasoning-only add-on s", "Shared train+eval s (both arms pay)",
                 "n rounds averaged"], True)
    for i, lab in enumerate(sheet_order):
        b = R[lab]["mean"]
        t, o = lab.split(" / ")
        row = r2 + 2 + i
        if b is None:
            put(row, [t, o] + [UNM] * 12 +
                ["BUDGET=1 by instruction: only one round exists, so no mean and no spread"])
            continue
        put(row, [t, o, f(b["baseline_s"], 1, b["baseline_s_sd"]),
                  f(b["diseil_s"], 1, b["diseil_s_sd"]), f(b["vlm_tokens"], 0,
                  b["vlm_tokens_sd"]), f(b["llm_tokens"], 0, b["llm_tokens_sd"]),
                  b["overhead_x"], b["kag_tokens"],
                  f(b["reasoning_tokens"], 0, b["reasoning_tokens_sd"]),
                  b["reasoning_narrow_s"], b["reasoning_wide_s"], b["baseline_gate_s"],
                  b["reasoning_wide_addon_s"], b["train_eval_shared_s"], b["n_rounds"]])

    r3 = r2 + 8
    notes = [
        "READ BEFORE QUOTING ANY NUMBER",
        "Baseline arm = SafeDAgger (ablation=safe), matching the three RoboSuite D5 jobs.",
        "P1 = the FIRST round of the run: round 0 for the distil module (Door/Wipe/GridWorld, "
        "0-indexed) and round 1 for the fork's 1-indexed Push-T loop. It is the WORST case "
        "(weakest policy -> most failures -> most analysis), so every DISTIL figure under P1 "
        "is an UPPER BOUND on steady-state per-round cost, not an average.",
        "P5 = mean ± sample SD over the LLM-active rounds (0..4) of the BUDGET=5 RoboSuite "
        "runs, matched arm-for-arm against the baseline's same-indexed rounds. The terminal "
        "eval-only round 5 is excluded from BOTH arms. Push-T and GridWorld have no P5: they "
        "were run at BUDGET=1 by the author's explicit instruction.",
        "Overhead × is close to 1 on the RoboSuite tasks and UNDERSTATES the picture: a round's "
        "wall-clock is dominated by the from-scratch policy RETRAIN plus the 100-episode "
        "held-out EVAL, and BOTH arms pay all of it (the 'Shared train+eval s' column). The "
        "honest characterisation of DISTIL's cost is the reasoning-only add-on, in seconds and "
        "in tokens, not the ratio.",
        "reasoning-only NARROW = clustering + VLM + reasoning LLM + prescription + feasibility. "
        "It EXCLUDES the failure-screening rollout. The baseline does none of this work.",
        "reasoning-only WIDE = NARROW + the failure-screening rollout, i.e. every DISTIL-specific "
        "stage (the whole round minus the shared train+eval). The add-on column is "
        "WIDE minus the baseline's gate rollout.",
        "Reasoning LLM tokens = hidden thinking tokens billed in the completion, ALL stages. "
        "Push-T's 7,214 is its all-stage total; the 5,612 quoted in the Push-T source doc is the "
        "Reasoning stage alone and is NOT comparable with the other rows.",
        "Token counts are NOT comparable across rows: different backends (OpenRouter qwen/qwen3-32b "
        "for Door/Wipe/GridWorld; local vLLM qwen3-32b + qwen3-vl-32b for Push-T), different "
        "prompt sizes, different tasks.",
        "Single seed (seed 1) throughout. The SD in Block 2 is the round-to-round spread WITHIN "
        "one run, not a seed-to-seed spread.",
    ]
    put(r3, ["CAVEATS"], True)
    for i, n in enumerate(notes):
        ws.cell(r3 + 1 + i, 1).value = n

    r4 = r3 + 1 + len(notes) + 1
    put(r4, ["PROVENANCE"], True)
    prov = [f"{lab}: DISTIL job {jobs[lab][0]} / SafeDAgger job {jobs[lab][1]}"
            for lab in sheet_order]
    prov += [
        "Stage spans reconstructed from run.log timestamps and CHECKED against "
        "result.json history[].sec (agreement within 1 s on all 24+ completed rounds).",
        "Scripts: distil/scripts/parse_d5_stages.py -> merge_d5_compute.py -> write_d5_outputs.py",
        f"Generated {datetime.now():%Y-%m-%d %H:%M}",
    ]
    for i, p in enumerate(prov):
        ws.cell(r4 + 1 + i, 1).value = p

    wb.save(XLSX)
    wb2 = openpyxl.load_workbook(XLSX)
    after = list(wb2.sheetnames)
    assert after == before, f"SHEETS LOST!\nbefore={before}\nafter={after}"
    print(f"[xlsx] D5_Compute written; {len(after)} sheets intact: {after}")
    print(f"[write] {BUILD/'d5_compute.csv'}")

    # ---------------------------------------------------------------- MD
    write_md(R, jobs, order)
    print(f"[write] {BUILD/'d5_compute.md'}")
    return R, jobs, order


def write_md(R, jobs, order):
    L = []
    A = L.append
    A("# D5_Compute — wall-clock and token cost per round (complete five-setting matrix)")
    A("")
    A("Baseline arm = **SafeDAgger** (`ablation=safe`) in all five settings, matching the "
      "three RoboSuite D5 jobs. Method = **DISEIL** (`p4_subtask` / `p4_top3_rotate` are "
      "code identifiers only). Seed 1 throughout.")
    A("")
    A("Every number below traces to a logged event in a run that executed on this cluster. "
      "Nothing is extrapolated or hand-filled. Quantities that could not be measured are "
      f"written `{UNM}` with the reason.")
    A("")
    A("---")
    A("")
    A("## 0. The one thing to take away")
    A("")
    A("**A round's wall-clock is dominated by the from-scratch policy RETRAIN plus the "
      "100-episode held-out EVAL — and BOTH arms pay all of it.** On the three RoboSuite "
      "settings that shared cost is 548–1 491 s of the round. The consequence:")
    A("")
    A("* the raw **Overhead ×** on RoboSuite is **1.13–1.53×**, i.e. close to 1, and it "
      "**UNDERSTATES** DISEIL's cost, because the large shared denominator dilutes it;")
    A("* the quantity that actually characterises DISEIL is the **reasoning-only add-on** — "
      "the seconds and the tokens the baseline does **not** spend.")
    A("")
    A("Report both. Neither alone is honest.")
    A("")
    A("---")
    A("")
    A("## 1. Protocols — do not mix these")
    A("")
    A("| | **P1 — first round** | **P5 — mean of the LLM-active rounds** |")
    A("|---|---|---|")
    A("| Which rounds | the run's **first** round: round **0** for the distil module "
      "(Door/Wipe/GridWorld, 0-indexed), round **1** for the fork's 1-indexed Push-T loop | "
      "rounds **0–4** (the five LLM-active rounds) of the BUDGET=5 runs, matched "
      "arm-for-arm against the baseline's same-indexed rounds |")
    A("| Available for | **all five settings** | **the three RoboSuite settings only** |")
    A("| Why | Push-T and GridWorld were run at **BUDGET=1** (the author's explicit "
      "instruction), so the first round is the *only* round they have. P1 is therefore the "
      "only apples-to-apples comparison across all five. | Push-T/GridWorld: "
      f"**{UNM}** — at BUDGET=1 exactly one round exists, so there is no multi-round mean "
      "and no spread. |")
    A("| What it means | **Worst case / UPPER BOUND.** The first round has the weakest "
      "policy → most failures → biggest cluster set → most VLM+LLM calls → highest "
      "reasoning cost. | **Best estimate of steady-state cost.** ± is the sample SD of the "
      "round-to-round spread *within* one run (**not** a seed-to-seed spread). |")
    A("")
    A("Round 5 of the BUDGET=5 runs is a terminal, eval-only round (budget exhausted, no "
      "LLM); it is excluded from **both** arms under P5.")
    A("")
    A("---")
    A("")
    A("## 2. THE MATRIX — Protocol P1 (first round; all five settings)")
    A("")
    A("| Task / Obs | Baseline s/round | DISEIL s/round | Overhead × | VLM tok/round | "
      "LLM tok/round | KAG tok contribution | Reasoning-LLM tok/round |")
    A("|---|---|---|---|---|---|---|---|")
    for lab in order:
        b = R[lab]["first"]
        A(f"| **{lab}** | {f(b['baseline_s'],1)} | {f(b['diseil_s'],1)} | "
          f"**{b['overhead_x']}×** | {f(b['vlm_tokens'])} | {f(b['llm_tokens'])} | "
          f"{f(b['kag_tokens'])} | {f(b['reasoning_tokens'])} |")
    A("")
    A("### 2b. The same rounds, decomposed into shared vs DISEIL-specific seconds")
    A("")
    A("| Task / Obs | Shared train+eval (both arms) | DISEIL screen | DISEIL analysis+"
      "prescription | Baseline gate rollout | **Reasoning-only add-on** |")
    A("|---|---|---|---|---|---|")
    for lab in order:
        b = R[lab]["first"]
        A(f"| **{lab}** | {f(b['train_eval_shared_s'],1)} s | {f(b['screen_s'],1)} s | "
          f"{f(b['reasoning_narrow_s'],1)} s | {f(b['baseline_gate_s'],1)} s | "
          f"**+{f(b['reasoning_wide_addon_s'],1)} s** |")
    A("")
    A("* **reasoning-only NARROW** = clustering + VLM + reasoning LLM + prescription + "
      "feasibility (the *analysis+prescription* column). It **excludes** the "
      "failure-screening rollout. The baseline does **none** of this work.")
    A("* **reasoning-only WIDE** = NARROW + the failure-screening rollout = every "
      "DISEIL-specific stage (the whole round minus the shared train+eval).")
    A("* **Reasoning-only add-on** = WIDE − the baseline's gate rollout. This is the "
      "honest \"what DISEIL costs you\" number.")
    A("")
    A("⚠️ **The two source documents defined `reasoning-only` differently and this merge "
      "fixes that.** The Push-T/GridWorld source doc reported Push-T's 491.7 s as "
      "NARROW (screen excluded) but GridWorld's 65.5 s as WIDE (screen included). Both "
      "definitions are computed consistently for all five settings here:")
    A("")
    A("| Task / Obs | reasoning-only **narrow** s | reasoning-only **wide** s |")
    A("|---|---|---|")
    for lab in order:
        b = R[lab]["first"]
        A(f"| {lab} | {f(b['reasoning_narrow_s'],1)} | {f(b['reasoning_wide_s'],1)} |")
    A("")
    A("---")
    A("")
    A("## 3. THE MATRIX — Protocol P5 (mean ± SD over the LLM-active rounds)")
    A("")
    A("The **n rounds** column states exactly how many rounds each mean is taken over. "
      "It is 5 for a completed BUDGET=5 run; a smaller number means the run was still in "
      "flight when this matrix was generated and the mean is over the rounds that had "
      "actually completed. No round is ever extrapolated.")
    A("")
    A("| Task / Obs | Baseline s/round | DISEIL s/round | Overhead × | VLM tok/round | "
      "LLM tok/round | KAG tok contribution | Reasoning-LLM tok/round | n rounds |")
    A("|---|---|---|---|---|---|---|---|---|")
    for lab in order:
        b = R[lab]["mean"]
        if b is None:
            A(f"| **{lab}** | {UNM} | {UNM} | {UNM} | {UNM} | {UNM} | {UNM} | {UNM} | — |")
            continue
        A(f"| **{lab}** | {f(b['baseline_s'],1,b['baseline_s_sd'])} | "
          f"{f(b['diseil_s'],1,b['diseil_s_sd'])} | **{b['overhead_x']}×** | "
          f"{f(b['vlm_tokens'],0,b['vlm_tokens_sd'])} | "
          f"{f(b['llm_tokens'],0,b['llm_tokens_sd'])} | {f(b['kag_tokens'])} | "
          f"{f(b['reasoning_tokens'],0,b['reasoning_tokens_sd'])} | {b['n_rounds']} |")
    A("")
    A(f"Push-T and GridWorld are **{UNM}** under P5 for one reason and one reason only: "
      "they were run at **BUDGET=1** on the author's explicit instruction, so exactly one "
      "round exists. A mean and a spread over 5 rounds cannot be computed from 1 round, "
      "and will not be invented here.")
    A("")
    A("### 3b. P5, decomposed")
    A("")
    A("| Task / Obs | Shared train+eval | DISEIL screen | DISEIL analysis+prescription | "
      "Baseline gate | **Reasoning-only add-on** |")
    A("|---|---|---|---|---|---|")
    for lab in order:
        b = R[lab]["mean"]
        if b is None:
            A(f"| {lab} | {UNM} | {UNM} | {UNM} | {UNM} | {UNM} |")
            continue
        A(f"| **{lab}** | {f(b['train_eval_shared_s'],1)} s | {f(b['screen_s'],1)} s | "
          f"{f(b['reasoning_narrow_s'],1)} s | {f(b['baseline_gate_s'],1)} s | "
          f"**+{f(b['reasoning_wide_addon_s'],1)} s** |")
    A("")
    A("---")
    A("")
    A("## 4. Why Push-T and GridWorld show a much larger Overhead × than Door and Wipe")
    A("")
    A("It is **not** that their reasoning is more expensive relative to the work done. It "
      "is that their **shared denominator is small** and, on Push-T, that the baseline's "
      "rollout is atypically tiny at BUDGET=1:")
    A("")
    A("* **SafeDAgger's loop is `while interventions < budget`.** At BUDGET=1 it stops at "
      "its **first** intervened episode — 1 episode, 6.3 s on Push-T. DISEIL meanwhile "
      "screens a fixed 60 episodes (746.8 s). That asymmetry, not the retrain, is what "
      "produces the 2.75×. At a realistic budget the baseline would roll out many episodes "
      "per retrain and the ratio would **fall**.")
    A("* On Door/Wipe the shared train+eval is 548–1 491 s per round — large enough to "
      "swamp a 234–704 s reasoning add-on, which is why those ratios sit at 1.13–1.53×.")
    A("")
    A("Both effects push the ratio around without changing the underlying reasoning cost. "
      "**This is precisely why the ratio is the wrong headline and the add-on is the right "
      "one.**")
    A("")
    A("---")
    A("")
    A("## 5. Provenance — job IDs and exact commands")
    A("")
    A("| Setting | DISEIL job | SafeDAgger job | Budget | Backend |")
    A("|---|---|---|---|---|")
    backend = {
        "Door / state": "OpenRouter `qwen/qwen3-32b` + VLM",
        "Door / image": "OpenRouter `qwen/qwen3-32b` + VLM",
        "Wipe / image": "OpenRouter `qwen/qwen3-32b` + VLM",
        "Push-T / image": "local vLLM `qwen3-32b` + `qwen3-vl-32b`",
        "GridWorld 5x5 / image": "OpenRouter `qwen/qwen3-32b` + VLM",
    }
    bud = {"Push-T / image": 1, "GridWorld 5x5 / image": 1}
    for lab in order:
        jd, jb = jobs[lab]
        A(f"| {lab} | **{jd}** | **{jb}** | {bud.get(lab,5)} | {backend[lab]} |")
    A("")
    A("```bash")
    A("cd /weka/s226137394/DmNfull")
    A("")
    A("# ---- the three RoboSuite settings (BUDGET=5) : jobs 110355-110360 ----")
    A("# (already launched by the main build; DO NOT relaunch)")
    A("sbatch --job-name=d5_Door_state_full \\")
    A("  --export=ALL,TASK=Door,MODALITY=state,ABLATION=full,SEED=1,BUDGET=5 \\")
    A("  distil/scripts/run_d5.sbatch                       # -> 110355")
    A("sbatch --job-name=d5_Door_state_safe \\")
    A("  --export=ALL,TASK=Door,MODALITY=state,ABLATION=safe,SEED=1,BUDGET=5 \\")
    A("  distil/scripts/run_d5.sbatch                       # -> 110356")
    A("#   ... likewise Door/image -> 110357/110358, Wipe/image -> 110359/110360")
    A("")
    A("# ---- Push-T / image (BUDGET=1) : jobs 110375 / 110376 ----")
    A("sbatch --job-name=d5_PushT_image_full \\")
    A("  --export=ALL,METHODS=p4_subtask,SEED=1,RUN_ID=511 \\")
    A("  distil/scripts/run_pusht_d5.sbatch                 # -> 110375")
    A("sbatch --job-name=d5_PushT_image_safe --gpus-per-node=1 \\")
    A("  --constraint=\"gpu-h100|gpu-h200\" \\")
    A("  --export=ALL,METHODS=safe_dagger,SEED=1,RUN_ID=511 \\")
    A("  distil/scripts/run_pusht_d5.sbatch                 # -> 110376")
    A("")
    A("# ---- GridWorld 5x5 / image (BUDGET=1) : jobs 110384 / 110385 ----")
    A("# CONDA_ENV=diffdagger is REQUIRED (no `distil` env on this cluster; without the")
    A("# override `python` resolves to an interpreter with no torch -- this killed 110377/8)")
    A("sbatch --job-name=d5_GridWorld_image_full \\")
    A("  --export=ALL,MODALITY=image,ABLATION=full,SEED=1,BUDGET=1,CONDA_ENV=diffdagger \\")
    A("  distil/scripts/run_gridworld_d5.sbatch             # -> 110384")
    A("sbatch --job-name=d5_GridWorld_image_safe \\")
    A("  --export=ALL,MODALITY=image,ABLATION=safe,SEED=1,BUDGET=1,CONDA_ENV=diffdagger \\")
    A("  distil/scripts/run_gridworld_d5.sbatch             # -> 110385")
    A("")
    A("# ---- rebuild this matrix from the runs (read-only over the runs) ----")
    A("python3 distil/scripts/parse_d5_stages.py    # run.log -> d5_stage_spans.json")
    A("python3 distil/scripts/merge_d5_compute.py   # + pusht/gridworld -> d5_merged.json")
    A("python3 distil/scripts/write_d5_outputs.py   # -> d5_compute.{md,csv} + the workbook")
    A("```")
    A("")
    A("### How the seconds were measured")
    A("")
    A("The RoboSuite runs log no per-stage spans, so stage boundaries were reconstructed "
      "from the `[HH:MM:SS]` timestamps that `distil/run.py` already prints "
      "(`[train]` → `[calibrate]` → `[eval]` → `[screen]` → `[distil-llm] … analysed` → "
      "`[collect a0]`). The reconstruction is **checked** against the independently "
      "recorded `result.json` `history[].sec`: it agrees within **1 s on every one of the "
      "completed rounds**. Push-T and GridWorld carry explicit span events in their own "
      "telemetry side-files and needed no reconstruction.")
    A("")
    A("Sources consumed:")
    A("")
    A("* RoboSuite: `distil/results/_compute/{Door,Wipe}/{state,image}/{full,safe}/seed1/"
      "{run.log,result.json,telemetry/round_*.jsonl}`")
    A("* Push-T: `…/pool_rl_robo/results/PushT-v1/run_511/{p4_subtask,safe_dagger}/results/"
      "telemetry/d5_events.jsonl`")
    A("* GridWorld: `distil/results/_compute/GridWorld/image/{full,safe}/seed1/telemetry/"
      "compute.jsonl`")
    A("* KAG tokens/call: `paper_aaai2027/COC_REPORT/build/kag_tokens.json`")
    A("* Merged intermediates: `distil/results/_compute/{d5_stage_spans,d5_merged}.json`")
    A("")
    A("---")
    A("")
    A("## 6. CAVEATS — read before quoting any number")
    A("")
    A("**C1 — Single seed (seed 1) everywhere.** The ± in §3 is the round-to-round spread "
      "*within one run*. It is **not** a seed-to-seed confidence interval. No cell in this "
      "matrix has cross-seed variance.")
    A("")
    A("**C2 — P1 is an UPPER BOUND, not an average.** The first round has the weakest "
      "policy and therefore the most failures. Failure count drives everything "
      "DISEIL-specific: episodes analysed → VLM calls → reasoning-LLM calls → "
      "KAG-carrying prompts → tokens and reasoning seconds. Later rounds cost less "
      "(Door/state screen failures fall 19 → 12 → 15 → 9 → 6 across rounds 0–4, and the "
      "DISEIL round falls 1 054 s → 587 s). Compare §2 with §3 to see the size of the "
      "effect.")
    A("")
    A("**C3 — Round 0 carries a double-length initial train, for BOTH arms.** The distil "
      "module uses `initial_train_steps=8000` in round 0 and `round_train_steps=4000` "
      "thereafter (Door/state DISEIL train: 250 s in round 0, then ~131–137 s). Since "
      "that inflated train sits in the **denominator** of Overhead ×, the P1 ratio is "
      "computed against an inflated *shared* cost and is therefore **conservative** "
      "(too low). The reasoning-only add-on is immune to this and is the robust figure.")
    A("")
    A("**C4 — The Overhead × is not a measure of the reasoning cost.** See §0 and §4. On "
      "RoboSuite it is diluted by the large shared train+eval; on Push-T it is inflated by "
      "SafeDAgger's 1-episode rollout at BUDGET=1. It moves for reasons that have nothing "
      "to do with how expensive DISEIL's reasoning is.")
    A("")
    A("**C5 — Token counts are NOT comparable across rows.** Door/Wipe/GridWorld run on "
      "**OpenRouter** (`qwen/qwen3-32b` + VLM) and read hidden-thinking tokens directly "
      "from `usage.completion_tokens_details.reasoning_tokens`. Push-T runs on **local "
      "vLLM** (`qwen3-32b` + `qwen3-vl-32b`) and *recovers* hidden-thinking tokens as "
      "`completion_tokens − tokens(visible text)` with the serving tokenizer, because the "
      "proxy strips `<think>…</think>` from the text while vLLM still bills those tokens. "
      "Both are measured, but they are different instruments over different models, "
      "prompts and tasks. Push-T's much larger counts (82 k vs ~10 k) reflect 20 LLM calls "
      "against 7, not a more expensive method.")
    A("")
    A("**C6 — 'Reasoning-LLM tokens/round' is the ALL-STAGE hidden-thinking total.** For "
      "the RoboSuite/GridWorld rows that is `tokens.reasoning` = analysis + decision (the "
      "VLM emits ~0). For Push-T the comparable all-stage total is **7 214** "
      "(vlm 9 + reasoning-stage 5 612 + plain aggregator 1 593). The Push-T source doc's "
      "headline **5 612** is the *Reasoning stage alone* and must not be placed in the "
      "same column as the other four rows — this merge uses 7 214.")
    A("")
    A("**C7 — KAG token contribution is measured per call, then multiplied by that run's "
      "own KAG-carrying call count** (`kag_calls` in the telemetry; VLM prompts carry no "
      "KAG). Tokens per KAG block, by paired prompt-token diff at `max_tokens=1` against "
      "the serving tokenizer: **Door 800**, **Wipe 598**, **GridWorld 865**, "
      "**Push-T 1 433**. KAG is a large share of the prompt budget — 41–42 % on Door, "
      "35 % on Wipe, **54 % on GridWorld**, 17 % on Push-T.")
    A("")
    A("**C8 — Push-T's LLM calls run CONCURRENTLY.** Its 20 calls sum to 1 040.9 s of "
      "*serving* time but their wall-clock union is 487.9 s. All seconds in this document "
      "are **wall-clock**, never summed call latencies.")
    A("")
    A("**C9 — Bootstrap demo collection is excluded from every per-round total, in both "
      "arms.** Both Push-T arms additionally reuse an identical bootstrap checkpoint "
      "(`P4_REUSE_INIT_CKPT`), so no bootstrap training is re-paid and the arms start from "
      "the same policy.")
    A("")
    A(f"**C10 — Nothing in this matrix is `{UNM}` except the P5 cells for Push-T and "
      "GridWorld**, and those are unmeasurable for the stated structural reason "
      "(BUDGET=1 → one round → no mean, no spread). No timing and no token count anywhere "
      "in this document was estimated, extrapolated or invented.")
    A("")
    A("---")
    A("")
    A("## 7. Instrumentation note (shared repo)")
    A("")
    A("All D5 instrumentation is **additive and non-breaking**. "
      "`pool_rl_robo/telemetry_d5.py` is inert unless `D5_TELEMETRY=1` and only wraps "
      "functions call-through, appending to a new side-file. `distil/compute_log.py` only "
      "appends `telemetry/compute.jsonl` and does not touch `result.json`'s schema. The "
      "three scripts added for this merge "
      "(`parse_d5_stages.py`, `merge_d5_compute.py`, `write_d5_outputs.py`) are "
      "**read-only** over the runs — they parse `run.log` / `result.json` / telemetry and "
      "write new side-files. No existing behaviour, default, threshold, control-flow branch "
      "or output schema was changed, and no running job was touched.")
    A("")
    A(f"_Generated {datetime.now():%Y-%m-%d %H:%M} from "
      "`distil/results/_compute/d5_merged.json`._")
    (BUILD / "d5_compute.md").write_text("\n".join(L) + "\n")


if __name__ == "__main__":
    main()
